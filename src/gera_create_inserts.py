import os
import argparse
import csv
import random
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from enum import Enum
from typing import Dict, List, Optional, Sequence, Set, Tuple

import pandas as pd
import tkinter as tk
from tkinter import filedialog
import chardet
import unidecode
import datetime as dt_module
from executa_arquivo_sql import execute_sql_file

# Limites da amostragem para inferência de tipos
MAX_LINHAS_AMOSTRA = 200
AMOSTRA_INICIO = 50
AMOSTRA_MEIO = 50
AMOSTRA_FIM = 50
AMOSTRA_ALEATORIA = 50
LIMITE_CARACTERES_TEXTO = 50
VARCHAR_TAMANHO_MINIMO = 1
VARCHAR_MARGEM_SEGURANCA = 2
INT32_MIN = -2_147_483_648
INT32_MAX = 2_147_483_647

VALORES_BOOLEANOS = {
    'true', 'false', 't', 'f', 'sim', 'nao', 'não', 'yes', 'no', 'y', 'n', '1', '0',
}

PADROES_DATA = (
    re.compile(r'^\d{4}-\d{2}-\d{2}$'),
    re.compile(r'^\d{2}/\d{2}/\d{4}$'),
    re.compile(r'^\d{2}-\d{2}-\d{4}$'),
    re.compile(r'^\d{4}/\d{2}/\d{2}$'),
)

PADROES_TIMESTAMP = (
    re.compile(r'^\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}:\d{2}(?:\.\d+)?$'),
    re.compile(r'^\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2}$'),
    re.compile(r'^\d{2}-\d{2}-\d{4}\s+\d{2}:\d{2}:\d{2}$'),
    re.compile(r'^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}(?:\.\d+)?(?:Z|[+-]\d{2}:?\d{2})?$'),
)


class CategoriaValor(Enum):
    INTEIRO = 'inteiro'
    DECIMAL = 'decimal'
    DATA = 'data'
    TIMESTAMP = 'timestamp'
    BOOLEANO = 'booleano'
    TEXTO = 'texto'


def format_file_name(file_name):
    formatted_name = unidecode.unidecode(str(file_name))
    formatted_name = re.sub(r'\s+', ' ', formatted_name).strip()
    formatted_name = formatted_name.replace(' ', '_')
    formatted_name = re.sub(r'_{2,}', '_', formatted_name)
    formatted_name = formatted_name.strip('_')
    return formatted_name


def sanitizar_identificador(identificador: str) -> str:
    """Remove caracteres inválidos e normaliza espaços/underscores em identificadores SQL."""
    nome = format_file_name(identificador)
    nome = re.sub(r'[^a-zA-Z0-9_]', '', nome)
    nome = re.sub(r'_{2,}', '_', nome).strip('_')
    return nome


def normalizar_nome_coluna(nome_coluna: str) -> str:
    nome = sanitizar_identificador(str(nome_coluna)).lower()
    if not nome:
        nome = 'coluna'
    if nome[0].isdigit():
        nome = f'col_{nome}'
    return nome


def detectar_codificacao(arquivo):
    with open(arquivo, 'rb') as f:
        dados = f.read()
        resultado = chardet.detect(dados)
        return resultado['encoding']


def resolver_codificacao(codific):
    """Converte atalho numérico ou nome de codificação para o valor usado na leitura do CSV."""
    if codific == '1':
        return 'utf-8'
    if codific == '2':
        return 'iso-8859-1'
    if codific == '3':
        return 'windows-1252'
    return codific


def perguntar_codificacao():
    """Solicita ao usuário a codificação para leitura de arquivos CSV."""
    print("\nCodificação para leitura de arquivos CSV:")
    print("  1 - UTF-8 (padrão)")
    print("  2 - ISO-8859-1 (Latin-1)")
    print("  3 - Windows-1252")
    print("  ou informe outro nome de codificação suportado pelo Python")
    codific = input("Informe a opção (Enter para UTF-8): ").strip()
    if not codific:
        codific = '1'
    codificacao = resolver_codificacao(codific)
    print(f'Codificação selecionada: {codificacao}')
    return codificacao


def perguntar_modo_tipos():
    """Solicita ao usuário o modo de definição dos tipos de colunas no CREATE TABLE."""
    print("\nModo de tipos de dados para CREATE TABLE:")
    print("  1 - Todos os campos como text (padrão)")
    print("  2 - Inferir tipos automaticamente")
    opcao = input("Informe a opção (Enter para text): ").strip()
    inferir = opcao == '2'
    modo = 'inferência automática' if inferir else 'text'
    print(f'Modo selecionado: {modo}')
    return inferir


def perguntar_validacao_completa():
    """Solicita se a inferência deve validar todas as colunas em todas as linhas."""
    print("\nValidação de tipos inferidos:")
    print("  1 - Rápida: amostra distribuída + ajuste de VARCHAR no arquivo completo (padrão)")
    print("  2 - Completa: analisar todas as colunas em todas as linhas (pode ser mais lenta)")
    opcao = input("Informe a opção (Enter para rápida): ").strip()
    completa = opcao == '2'
    modo = 'validação completa' if completa else 'validação rápida'
    print(f'Validação selecionada: {modo}')
    return completa


def perguntar_schema(padrao='dados_gts'):
    """Solicita ao usuário o schema PostgreSQL de destino."""
    print("\nSchema PostgreSQL para gerar o SQL:")
    nome = input(f"Informe o nome do schema (Enter para {padrao}): ").strip()
    if not nome:
        nome = padrao
    print(f'Schema selecionado: {nome}')
    return nome

def tratar_dado(dado):
    if isinstance(dado, str):
        dado_tratado = dado.replace("_x000D_", "").replace("'", "''").replace(" ", "")
    elif isinstance(dado, float):
        if dado.is_integer():
            dado_tratado = str(int(dado))
        else:
            dado_tratado = str(dado)
    else:
        dado_tratado = str(dado)
    return dado_tratado


def valor_vazio(valor) -> bool:
    if valor is None or pd.isna(valor):
        return True
    if isinstance(valor, str) and not valor.strip():
        return True
    return False


def normalizar_colunas(colunas: Sequence[str]) -> List[str]:
    nomes_normalizados = []
    contagem = {}
    for coluna in colunas:
        nome = normalizar_nome_coluna(coluna)
        if nome in contagem:
            contagem[nome] += 1
            nome = f'{nome}_{contagem[nome]}'
        else:
            contagem[nome] = 0
        nomes_normalizados.append(nome)
    return nomes_normalizados


def citar_identificador(identificador: str) -> str:
    """Coloca o identificador entre aspas duplas (palavras reservadas do PostgreSQL)."""
    return f'"{identificador}"'


def selecionar_indices_amostra(total_linhas: int, max_linhas: int = MAX_LINHAS_AMOSTRA) -> List[int]:
    if total_linhas <= 0:
        return []
    if total_linhas <= max_linhas:
        return list(range(total_linhas))

    indices: Set[int] = set()

    indices.update(range(min(AMOSTRA_INICIO, total_linhas)))

    inicio_meio = max(0, (total_linhas // 2) - (AMOSTRA_MEIO // 2))
    indices.update(range(inicio_meio, min(inicio_meio + AMOSTRA_MEIO, total_linhas)))

    indices.update(range(max(0, total_linhas - AMOSTRA_FIM), total_linhas))

    restante = sorted(set(range(total_linhas)) - indices)
    if restante:
        quantidade_aleatoria = min(AMOSTRA_ALEATORIA, len(restante))
        indices.update(random.sample(restante, quantidade_aleatoria))

    return sorted(indices)


def contar_linhas_csv(arquivo: str, codificacao: str) -> int:
    with open(arquivo, 'r', encoding=codificacao, newline='') as arquivo_csv:
        total = sum(1 for _ in arquivo_csv)
    return max(0, total - 1)


def ler_cabecalho_csv(arquivo: str, codificacao: str) -> List[str]:
    with open(arquivo, 'r', encoding=codificacao, newline='') as arquivo_csv:
        leitor = csv.reader(arquivo_csv, delimiter=';')
        return next(leitor)


def ler_linhas_csv_por_indices(
    arquivo: str,
    codificacao: str,
    indices: Sequence[int],
) -> pd.DataFrame:
    indices_desejados = set(indices)
    linhas: List[List[str]] = []
    cabecalho: List[str] = []

    with open(arquivo, 'r', encoding=codificacao, newline='') as arquivo_csv:
        leitor = csv.reader(arquivo_csv, delimiter=';')
        cabecalho = next(leitor)
        for indice, linha in enumerate(leitor):
            if indice in indices_desejados:
                linhas.append(linha)

    return pd.DataFrame(linhas, columns=cabecalho)


def contar_linhas_excel(arquivo: str) -> int:
    import openpyxl

    planilha = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
    try:
        aba = planilha.active
        total = 0
        for indice, _ in enumerate(aba.iter_rows(values_only=True)):
            if indice == 0:
                continue
            total += 1
        return total
    finally:
        planilha.close()


def ler_cabecalho_excel(arquivo: str) -> List[str]:
    import openpyxl

    planilha = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
    try:
        aba = planilha.active
        primeira_linha = next(aba.iter_rows(values_only=True))
        return [str(coluna) if coluna is not None else '' for coluna in primeira_linha]
    finally:
        planilha.close()


def ler_linhas_excel_por_indices(
    arquivo: str,
    indices: Sequence[int],
) -> pd.DataFrame:
    import openpyxl

    indices_desejados = set(indices)
    cabecalho: List[str] = []
    linhas: List[List[object]] = []

    planilha = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
    try:
        aba = planilha.active
        for indice_linha, linha in enumerate(aba.iter_rows(values_only=True)):
            if indice_linha == 0:
                cabecalho = [str(coluna) if coluna is not None else '' for coluna in linha]
                continue
            indice_dados = indice_linha - 1
            if indice_dados in indices_desejados:
                linhas.append(list(linha))
    finally:
        planilha.close()

    return pd.DataFrame(linhas, columns=cabecalho)


def ler_arquivo(nome_arquivo: str, codificacao: str = 'utf-8') -> pd.DataFrame:
    extensao = os.path.splitext(nome_arquivo)[1].lower()
    if extensao == '.xlsx':
        return pd.read_excel(nome_arquivo, engine='openpyxl')
    if extensao == '.csv':
        return pd.read_csv(nome_arquivo, encoding=codificacao, delimiter=';', low_memory=False)
    raise ValueError("Formato de arquivo não suportado.")


def ler_amostra_para_inferencia(
    nome_arquivo: str,
    codificacao: str,
) -> Tuple[pd.DataFrame, List[str]]:
    extensao = os.path.splitext(nome_arquivo)[1].lower()

    if extensao == '.csv':
        total_linhas = contar_linhas_csv(nome_arquivo, codificacao)
        cabecalho = ler_cabecalho_csv(nome_arquivo, codificacao)
        indices = selecionar_indices_amostra(total_linhas)
        if not indices:
            return pd.DataFrame(columns=cabecalho), normalizar_colunas(cabecalho)
        df_amostra = ler_linhas_csv_por_indices(nome_arquivo, codificacao, indices)
    elif extensao == '.xlsx':
        total_linhas = contar_linhas_excel(nome_arquivo)
        cabecalho = ler_cabecalho_excel(nome_arquivo)
        indices = selecionar_indices_amostra(total_linhas)
        if not indices:
            return pd.DataFrame(columns=cabecalho), normalizar_colunas(cabecalho)
        df_amostra = ler_linhas_excel_por_indices(nome_arquivo, indices)
    else:
        raise ValueError("Formato de arquivo não suportado.")

    colunas = normalizar_colunas(cabecalho)
    df_amostra.columns = colunas
    return df_amostra, colunas


def valor_para_texto(valor) -> str:
    if isinstance(valor, datetime):
        return valor.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(valor, date):
        return valor.strftime('%Y-%m-%d')
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def corresponde_inteiro(texto: str) -> bool:
    return bool(re.fullmatch(r'[+-]?\d+', texto))


def corresponde_decimal(texto: str) -> bool:
    return bool(
        re.fullmatch(r'[+-]?\d+[.,]\d+', texto)
        or re.fullmatch(r'[+-]?\d+\.\d+e[+-]?\d+', texto, flags=re.IGNORECASE)
    )


def corresponde_data(texto: str) -> bool:
    if any(padrao.match(texto) for padrao in PADROES_DATA):
        return _data_valida(texto)
    return False


def corresponde_timestamp(texto: str) -> bool:
    if any(padrao.match(texto) for padrao in PADROES_TIMESTAMP):
        return _timestamp_valido(texto)
    return False


def _data_valida(texto: str) -> bool:
    formatos = ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d')
    for formato in formatos:
        try:
            datetime.strptime(texto, formato)
            return True
        except ValueError:
            continue
    return False


def _timestamp_valido(texto: str) -> bool:
    candidatos = (
        texto,
        texto.replace('T', ' ').replace('Z', ''),
    )
    formatos = (
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d %H:%M:%S.%f',
        '%d/%m/%Y %H:%M:%S',
        '%d-%m-%Y %H:%M:%S',
        '%Y-%m-%dT%H:%M:%S',
        '%Y-%m-%dT%H:%M:%S.%f',
    )
    for candidato in candidatos:
        for formato in formatos:
            try:
                datetime.strptime(candidato, formato)
                return True
            except ValueError:
                continue
    return False


def corresponde_booleano(texto: str) -> bool:
    return texto.strip().lower() in VALORES_BOOLEANOS


def classificar_valor(valor) -> CategoriaValor:
    if isinstance(valor, bool):
        return CategoriaValor.BOOLEANO
    if isinstance(valor, datetime):
        if (
            valor.hour == 0
            and valor.minute == 0
            and valor.second == 0
            and valor.microsecond == 0
        ):
            return CategoriaValor.DATA
        return CategoriaValor.TIMESTAMP
    if isinstance(valor, date):
        return CategoriaValor.DATA
    if isinstance(valor, int) and not isinstance(valor, bool):
        return CategoriaValor.INTEIRO
    if isinstance(valor, float):
        if pd.isna(valor):
            return CategoriaValor.TEXTO
        if valor.is_integer():
            return CategoriaValor.INTEIRO
        return CategoriaValor.DECIMAL

    texto = valor_para_texto(valor)
    if not texto:
        return CategoriaValor.TEXTO

    if corresponde_booleano(texto):
        return CategoriaValor.BOOLEANO
    if corresponde_inteiro(texto):
        return CategoriaValor.INTEIRO
    if corresponde_decimal(texto):
        return CategoriaValor.DECIMAL
    if corresponde_timestamp(texto):
        return CategoriaValor.TIMESTAMP
    if corresponde_data(texto):
        return CategoriaValor.DATA
    return CategoriaValor.TEXTO


def coluna_exige_integer_por_nome(nome_coluna: str) -> bool:
    return nome_coluna == 'id' or nome_coluna.startswith('id_')


def valores_contem_texto_nao_numerico(valores: Sequence) -> bool:
    for valor in valores:
        if valor_vazio(valor):
            continue
        if isinstance(valor, (int, float)) and not isinstance(valor, bool):
            continue
        texto = valor_para_texto(valor)
        if not (corresponde_inteiro(texto) or corresponde_decimal(texto)):
            return True
    return False


def tipo_inteiro_por_magnitude(valores: Sequence) -> str:
    for valor in valores:
        if valor_vazio(valor):
            continue
        if isinstance(valor, bool):
            continue
        try:
            if isinstance(valor, int):
                numero = valor
            elif isinstance(valor, float):
                numero = int(valor)
            else:
                numero = int(valor_para_texto(valor))
        except (ValueError, TypeError, OverflowError):
            return inferir_varchar(valores)
        if numero < INT32_MIN or numero > INT32_MAX:
            return 'bigint'
    return 'integer'


def tamanho_maximo_amostra(valores: Sequence) -> int:
    tamanho_maximo = 0
    for valor in valores:
        if valor_vazio(valor):
            continue
        tamanho_maximo = max(tamanho_maximo, len(valor_para_texto(valor)))
    return tamanho_maximo


def coluna_excede_limite_texto(valores: Sequence) -> bool:
    """Verifica se algum valor ultrapassa o limite; interrompe na primeira ocorrência."""
    for valor in valores:
        if valor_vazio(valor):
            continue
        if len(valor_para_texto(valor)) > LIMITE_CARACTERES_TEXTO:
            return True
    return False


def valores_excedem_limite_texto(valores: Sequence) -> bool:
    return coluna_excede_limite_texto(valores)


def inferir_varchar(valores: Sequence, tamanho_maximo: Optional[int] = None) -> str:
    if tamanho_maximo is None:
        tamanho_maximo = tamanho_maximo_amostra(valores)
    return calcular_tipo_cadeia_por_tamanho(tamanho_maximo)


def arredondar_tamanho_varchar(tamanho: int) -> int:
    """Arredonda VARCHAR para a dezena superior; abaixo de 10 mantém o valor exato."""
    tamanho = max(tamanho, VARCHAR_TAMANHO_MINIMO)
    if tamanho < 10:
        return tamanho
    return ((tamanho + 9) // 10) * 10


def calcular_tipo_cadeia_por_tamanho(tamanho_maximo: int) -> str:
    if tamanho_maximo > LIMITE_CARACTERES_TEXTO:
        return 'text'
    tamanho = max(tamanho_maximo + VARCHAR_MARGEM_SEGURANCA, VARCHAR_TAMANHO_MINIMO)
    if tamanho > LIMITE_CARACTERES_TEXTO:
        return 'text'
    tamanho = arredondar_tamanho_varchar(tamanho)
    tamanho = min(tamanho, LIMITE_CARACTERES_TEXTO)
    return f'varchar({tamanho})'


def tamanho_maximo_coluna(df: pd.DataFrame, coluna: str) -> int:
    if coluna not in df.columns:
        return 0
    tamanho_maximo = 0
    for valor in df[coluna]:
        if valor_vazio(valor):
            continue
        tamanho = len(valor_para_texto(valor))
        if tamanho > LIMITE_CARACTERES_TEXTO:
            return LIMITE_CARACTERES_TEXTO + 1
        tamanho_maximo = max(tamanho_maximo, tamanho)
    return tamanho_maximo


def inferir_tipo_coluna_completo(df: pd.DataFrame, nome_coluna: str) -> str:
    """Inferência com saída antecipada: > limite de caracteres vira text imediatamente."""
    if nome_coluna not in df.columns:
        return inferir_varchar([])

    valores_validos = []
    for valor in df[nome_coluna]:
        if valor_vazio(valor):
            continue
        texto = valor_para_texto(valor)
        if len(texto) > LIMITE_CARACTERES_TEXTO:
            return 'text'
        valores_validos.append(valor)

    return inferir_tipo_coluna(nome_coluna, valores_validos)


def ajustar_tipos_pelo_dataset_completo(
    df: pd.DataFrame,
    tipos_colunas: Dict[str, str],
) -> Dict[str, str]:
    """Recalcula VARCHAR/text com base em todas as linhas, evitando subestimar pela amostra."""
    tipos_ajustados = dict(tipos_colunas)
    for coluna, tipo in tipos_colunas.items():
        if not tipo_eh_cadeia_curta(tipo):
            continue
        tamanho_maximo = tamanho_maximo_coluna(df, coluna)
        tipo_ajustado = calcular_tipo_cadeia_por_tamanho(tamanho_maximo)
        if tipo_ajustado != tipo:
            print(
                f'  - {coluna}: {tipo} -> {tipo_ajustado} '
                f'(tamanho máximo no arquivo: {tamanho_maximo})'
            )
        tipos_ajustados[coluna] = tipo_ajustado
    return tipos_ajustados


def validar_tipos_pelo_dataset_completo(
    df: pd.DataFrame,
    colunas: Sequence[str],
    tipos_referencia: Optional[Dict[str, str]] = None,
) -> Dict[str, str]:
    """Reinfere tipos de todas as colunas usando todas as linhas do arquivo."""
    total_colunas = len(colunas)
    tipos_validados = {}

    for indice, coluna in enumerate(colunas, start=1):
        tipo_novo = inferir_tipo_coluna_completo(df, coluna)
        tipo_anterior = tipos_referencia.get(coluna) if tipos_referencia else None
        if tipo_anterior and tipo_anterior != tipo_novo:
            print(f'  - {coluna}: {tipo_anterior} -> {tipo_novo}')
        tipos_validados[coluna] = tipo_novo

        if total_colunas >= 20 and (indice % 20 == 0 or indice == total_colunas):
            print(f'  Progresso da validação completa: {indice}/{total_colunas} colunas')

    return tipos_validados


def tipo_eh_cadeia_curta(tipo_coluna: str) -> bool:
    return bool(re.fullmatch(r'varchar\(\d+\)', tipo_coluna, flags=re.IGNORECASE))


def combinar_categorias(categorias: Set[CategoriaValor]) -> str:
    if not categorias or categorias == {CategoriaValor.TEXTO}:
        return 'varchar'

    if CategoriaValor.TEXTO in categorias:
        return 'varchar'

    if CategoriaValor.BOOLEANO in categorias:
        if categorias == {CategoriaValor.BOOLEANO}:
            return 'boolean'
        return 'varchar'

    if CategoriaValor.TIMESTAMP in categorias:
        if categorias <= {CategoriaValor.TIMESTAMP, CategoriaValor.DATA}:
            return 'timestamp'
        return 'varchar'

    if CategoriaValor.DATA in categorias:
        if categorias == {CategoriaValor.DATA}:
            return 'date'
        if categorias <= {CategoriaValor.DATA, CategoriaValor.TIMESTAMP}:
            return 'timestamp'
        return 'varchar'

    if CategoriaValor.DECIMAL in categorias:
        if categorias <= {CategoriaValor.DECIMAL, CategoriaValor.INTEIRO}:
            return 'numeric'
        return 'varchar'

    if categorias == {CategoriaValor.INTEIRO}:
        return 'integer'

    return 'varchar'


def inferir_tipo_coluna(nome_coluna: str, valores: Sequence) -> str:
    valores_validos = []
    for valor in valores:
        if valor_vazio(valor):
            continue
        texto = valor_para_texto(valor)
        if len(texto) > LIMITE_CARACTERES_TEXTO:
            return 'text'
        valores_validos.append(valor)

    if not valores_validos:
        return inferir_varchar(valores_validos)

    if coluna_exige_integer_por_nome(nome_coluna):
        if not valores_contem_texto_nao_numerico(valores_validos):
            tipo_inteiro = tipo_inteiro_por_magnitude(valores_validos)
            if tipo_inteiro in ('integer', 'bigint'):
                return tipo_inteiro

    categorias = {classificar_valor(valor) for valor in valores_validos}

    # Regra conservadora para booleanos: todos os valores precisam ser booleanos
    if CategoriaValor.BOOLEANO in categorias:
        textos = {valor_para_texto(valor).lower() for valor in valores_validos}
        if not textos.issubset(VALORES_BOOLEANOS):
            categorias.discard(CategoriaValor.BOOLEANO)
            if not categorias:
                return inferir_varchar(valores_validos)

    tipo = combinar_categorias(categorias)
    if tipo == 'integer':
        return tipo_inteiro_por_magnitude(valores_validos)
    if tipo == 'varchar':
        return inferir_varchar(valores_validos)
    return tipo


def inferir_tipos_colunas(df_amostra: pd.DataFrame, colunas: Sequence[str]) -> Dict[str, str]:
    tipos = {}
    for coluna in colunas:
        valores = df_amostra[coluna].tolist() if coluna in df_amostra.columns else []
        tipos[coluna] = inferir_tipo_coluna(coluna, valores)
    return tipos


def gerar_create_table(
    nome_schema: str,
    nome_tabela: str,
    colunas: Sequence[str],
    tipos_colunas: Optional[Dict[str, str]] = None,
) -> str:
    definicoes_colunas = []
    for coluna in colunas:
        tipo = tipos_colunas.get(coluna, 'text') if tipos_colunas else 'text'
        definicoes_colunas.append(f'{citar_identificador(coluna)} {tipo}')
    colunas_sql = ', '.join(definicoes_colunas)
    return (
        f'CREATE TABLE {nome_schema}.{citar_identificador(nome_tabela)} '
        f'(_id_origem serial, {colunas_sql});'
    )


def _converter_booleano_sql(valor) -> str:
    if isinstance(valor, bool):
        return 'true' if valor else 'false'
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        return 'true' if int(valor) == 1 else 'false'
    texto = valor_para_texto(valor).lower()
    if texto in {'true', 't', 'sim', 'yes', 'y', '1', 's'}:
        return 'true'
    if texto in {'false', 'f', 'nao', 'não', 'no', 'n', '0'}:
        return 'false'
    return f"'{tratar_dado(valor)}'"


def _converter_numero_sql(valor) -> str:
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        if isinstance(valor, float) and valor.is_integer():
            return str(int(valor))
        return str(valor)
    texto = valor_para_texto(valor).replace(',', '.')
    try:
        numero = Decimal(texto)
    except InvalidOperation:
        return f"'{tratar_dado(valor)}'"
    if numero == numero.to_integral_value():
        return str(int(numero))
    return format(numero.normalize(), 'f').rstrip('0').rstrip('.') or '0'


def _converter_data_sql(valor) -> str:
    if isinstance(valor, datetime):
        return f"'{valor.strftime('%Y-%m-%d')}'"
    if isinstance(valor, date):
        return f"'{valor.strftime('%Y-%m-%d')}'"
    texto = valor_para_texto(valor)
    formatos = ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d')
    for formato in formatos:
        try:
            data = datetime.strptime(texto, formato)
            return f"'{data.strftime('%Y-%m-%d')}'"
        except ValueError:
            continue
    return f"'{tratar_dado(valor)}'"


def _converter_timestamp_sql(valor) -> str:
    if isinstance(valor, datetime):
        return f"'{valor.strftime('%Y-%m-%d %H:%M:%S')}'"
    if isinstance(valor, date):
        return f"'{valor.strftime('%Y-%m-%d')}'"
    texto = valor_para_texto(valor).replace('T', ' ')
    formatos = (
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d %H:%M:%S.%f',
        '%d/%m/%Y %H:%M:%S',
        '%d-%m-%Y %H:%M:%S',
    )
    for formato in formatos:
        try:
            data_hora = datetime.strptime(texto, formato)
            return f"'{data_hora.strftime('%Y-%m-%d %H:%M:%S')}'"
        except ValueError:
            continue
    return f"'{tratar_dado(valor)}'"


def formatar_valor_insert(
    valor,
    tipo_coluna: str,
    usar_texto_para_todos: bool = False,
) -> str:
    if valor_vazio(valor):
        return 'NULL'

    if usar_texto_para_todos or tipo_coluna == 'text' or tipo_eh_cadeia_curta(tipo_coluna):
        return f"'{tratar_dado(valor)}'"

    if tipo_coluna == 'boolean':
        return _converter_booleano_sql(valor)

    if tipo_coluna in ('integer', 'bigint'):
        return _converter_numero_sql(valor)

    if tipo_coluna == 'numeric':
        return _converter_numero_sql(valor)

    if tipo_coluna == 'date':
        return _converter_data_sql(valor)

    if tipo_coluna == 'timestamp':
        return _converter_timestamp_sql(valor)

    return f"'{tratar_dado(valor)}'"


def gerar_insert_sql(
    nome_schema: str,
    nome_tabela: str,
    df: pd.DataFrame,
    colunas: Sequence[str],
    tipos_colunas: Optional[Dict[str, str]] = None,
    usar_texto_para_todos: bool = False,
) -> str:
    colunas_sql = ', '.join(citar_identificador(coluna) for coluna in colunas)
    insert_sql = (
        f'INSERT INTO {nome_schema}.{citar_identificador(nome_tabela)} '
        f'({colunas_sql}) VALUES\n'
    )

    total_linhas = df.shape[0]
    print("Total de linhas no DataFrame:", total_linhas)

    conta = 0
    for _, row in df.iterrows():
        conta += 1
        if conta % 25000 == 0:
            porcentagem = round((conta / total_linhas) * 100, 2)
            current_datetime = dt_module.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            print(
                'Data e hora:', current_datetime,
                '| Linhas processadas:', conta, ' - ', porcentagem, "%",
            )

        valores = []
        for coluna in colunas:
            tipo = tipos_colunas.get(coluna, 'text') if tipos_colunas else 'text'
            valores.append(
                formatar_valor_insert(
                    row[coluna],
                    tipo,
                    usar_texto_para_todos=usar_texto_para_todos,
                )
            )
        insert_sql += f"({', '.join(valores)}),\n"

    return insert_sql.rstrip(',\n') + ';'


def escrever_arquivo_sql(caminho_arquivo: str, conteudo: str) -> bool:
    """Grava o arquivo SQL, substituindo o conteúdo caso o arquivo já exista."""
    pasta_destino = os.path.dirname(caminho_arquivo)
    if pasta_destino:
        os.makedirs(pasta_destino, exist_ok=True)

    arquivo_existia = os.path.exists(caminho_arquivo)
    if arquivo_existia:
        print(f'Arquivo existente encontrado: {caminho_arquivo} — será substituído.')

    with open(caminho_arquivo, 'w', encoding='utf-8') as sql_file:
        sql_file.write(conteudo)

    return arquivo_existia


def gerar_sql(
    file_list,
    nome_schema,
    codificacao='utf-8',
    inferir_tipos=False,
    validacao_completa=False,
):
    if validacao_completa:
        inferir_tipos = True

    for nome_arquivo in file_list:
        try:
            print('Preparando dados do arquivo:', nome_arquivo)
            print('Codificação selecionada:', codificacao)
            if inferir_tipos:
                if validacao_completa:
                    print(
                        'Modo de tipos: inferência com validação completa '
                        '(todas as colunas e linhas — pode ser mais lenta)'
                    )
                else:
                    print(
                        'Modo de tipos: inferência rápida '
                        '(amostra + ajuste de VARCHAR no arquivo completo)'
                    )
            else:
                print('Modo de tipos: text (padrão)')

            tipos_colunas = None
            colunas = None

            if inferir_tipos and not validacao_completa:
                print('Coletando amostra para inferência de tipos...')
                df_amostra, colunas = ler_amostra_para_inferencia(nome_arquivo, codificacao)
                tipos_colunas = inferir_tipos_colunas(df_amostra, colunas)
                print('Tipos inferidos na amostra:')
                for coluna in colunas:
                    print(f'  - {coluna}: {tipos_colunas[coluna]}')

            print('Carregando dados na memória (df).')
            df = ler_arquivo(nome_arquivo, codificacao)

            print('Gerando dados para importa...')
            nome_arquivo_base = os.path.basename(nome_arquivo)
            nome_tabela = nome_arquivo_base.rsplit('.', 1)[0].lower()
            nome_tabela_formatado = format_file_name(nome_tabela)

            if colunas is None:
                colunas = normalizar_colunas(df.columns.tolist())
            df.columns = colunas

            if inferir_tipos and validacao_completa:
                print(
                    'Validação completa de tipos em todas as colunas '
                    '(pode ser mais lenta)...'
                )
                tipos_colunas = validar_tipos_pelo_dataset_completo(df, colunas, tipos_colunas)
                print('Tipos finais para CREATE TABLE:')
                for coluna in colunas:
                    print(f'  - {coluna}: {tipos_colunas[coluna]}')
            elif inferir_tipos and tipos_colunas is not None:
                print('Validando tamanhos VARCHAR no dataset completo...')
                tipos_colunas = ajustar_tipos_pelo_dataset_completo(df, tipos_colunas)
                print('Tipos finais para CREATE TABLE:')
                for coluna in colunas:
                    print(f'  - {coluna}: {tipos_colunas[coluna]}')

            print('Gerando CREATE TABLE...')
            create_sql = gerar_create_table(
                nome_schema,
                nome_tabela_formatado,
                colunas,
                tipos_colunas,
            )

            print('Gerando INSERT com os dados...')
            insert_sql = gerar_insert_sql(
                nome_schema,
                nome_tabela_formatado,
                df,
                colunas,
                tipos_colunas,
                usar_texto_para_todos=not inferir_tipos,
            )

            print("Escrevendo arquivo .sql")
            caminho_sql = f'sql/{nome_tabela}.sql'
            conteudo_sql = create_sql + '\n' + insert_sql
            arquivo_substituido = escrever_arquivo_sql(caminho_sql, conteudo_sql)

            if arquivo_substituido:
                print(f'O arquivo SQL "{nome_tabela}.sql" foi substituído com sucesso para {nome_arquivo}.')
            else:
                print(f'O arquivo SQL "{nome_tabela}.sql" foi criado com sucesso para {nome_arquivo}.')

            print('###' * 50)
            print('###' * 15, "Inserindo no Banco de dados", '###' * 15)
            print('    Executando script sql: ', nome_tabela + '.sql')
            if execute_sql_file('sql/' + nome_tabela + '.sql'):
                print('    Inserido com Sucesso!')
            else:
                print(f'    Erro ao inserir os dados do arquivo sql/{nome_tabela}.sql ')
            print('###' * 50)

        except Exception as e:
            print(f'Erro ao criar o arquivo SQL:', e)


def main():
    """Função principal para gerar CREATE e INSERTs"""
    parser = argparse.ArgumentParser(
        description="Gera scripts SQL (CREATE TABLE + INSERT) a partir de planilhas Excel ou CSV",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos de uso:
  python gera_create_inserts.py                              # Seleção interativa (UTF-8)
  python gera_create_inserts.py --codificacao utf-8          # CSV em UTF-8
  python gera_create_inserts.py -c 2                         # CSV em ISO-8859-1 (Latin-1)
  python gera_create_inserts.py -c windows-1252              # CSV em Windows-1252
  python gera_create_inserts.py --inferir-tipos              # Inferir tipos no CREATE TABLE
  python gera_create_inserts.py --validacao-completa         # Validação completa (mais lenta)
  python gera_create_inserts.py --perguntar-tipos            # Escolher modo de tipos interativamente
  python gera_create_inserts.py -s dados_gts                 # Schema sem prompt interativo

Codificações aceitas:
  1 ou utf-8          UTF-8 (padrão)
  2 ou iso-8859-1     ISO-8859-1 (Latin-1)
  3 ou windows-1252   Windows-1252
  ou qualquer nome de codificação suportado pelo Python
        """
    )

    parser.add_argument(
        '-c', '--codificacao',
        type=str,
        default='utf-8',
        metavar='CODIFICACAO',
        help='Codificação para leitura de arquivos CSV (padrão: utf-8). '
             'Atalhos: 1=utf-8, 2=iso-8859-1, 3=windows-1252'
    )

    parser.add_argument(
        '--perguntar-codificacao',
        action='store_true',
        help='Solicita interativamente a codificação dos arquivos CSV'
    )

    parser.add_argument(
        '--inferir-tipos',
        action='store_true',
        help='Infere automaticamente os tipos PostgreSQL das colunas no CREATE TABLE'
    )

    parser.add_argument(
        '--perguntar-tipos',
        action='store_true',
        help='Solicita interativamente o modo de tipos das colunas (text ou inferência)'
    )

    parser.add_argument(
        '--validacao-completa',
        action='store_true',
        help='Valida tipos inferidos em todas as colunas e linhas (mais lenta, mais segura)'
    )

    parser.add_argument(
        '--perguntar-validacao',
        action='store_true',
        help='Solicita interativamente o modo de validação dos tipos inferidos'
    )

    parser.add_argument(
        '-s', '--schema',
        type=str,
        default=None,
        metavar='SCHEMA',
        help='Schema PostgreSQL de destino (se omitido, será solicitado interativamente)'
    )

    args = parser.parse_args()
    if args.perguntar_codificacao:
        codificacao = perguntar_codificacao()
    else:
        codificacao = resolver_codificacao(args.codificacao)

    if args.perguntar_tipos:
        inferir_tipos = perguntar_modo_tipos()
    else:
        inferir_tipos = args.inferir_tipos or args.validacao_completa

    if inferir_tipos and args.perguntar_validacao:
        validacao_completa = perguntar_validacao_completa()
    else:
        validacao_completa = args.validacao_completa

    if args.schema:
        nome_schema = args.schema
    else:
        nome_schema = perguntar_schema()

    hora_inicio = dt_module.datetime.now()
    print("Hora inicial:", hora_inicio.strftime("%Y-%m-%d %H:%M:%S"))
    root = tk.Tk()
    root.withdraw()
    file_list = filedialog.askopenfilenames(
        title="Selecionar arquivos Excel ou CSV",
        filetypes=[
            ("Arquivos Excel", "*.xlsx"),
            ("Arquivos CSV", "*.csv"),
            ("Todos os arquivos", "*.*")
        ]
    )

    if file_list:
        print('Gerando dados para o schema:', nome_schema)

        gerar_sql(
            list(file_list),
            nome_schema,
            codificacao,
            inferir_tipos,
            validacao_completa,
        )

        print('Gerado com sucesso!!!')
        hora_fim = dt_module.datetime.now()
        duracao_total = hora_fim - hora_inicio

        print("Hora inicial:", hora_inicio.strftime("%Y-%m-%d %H:%M:%S"))
        print("Hora final:", hora_fim.strftime("%Y-%m-%d %H:%M:%S"))
        print("Tempo total de execução:", duracao_total)
    else:
        print("Nenhum arquivo selecionado.")


if __name__ == "__main__":
    main()
