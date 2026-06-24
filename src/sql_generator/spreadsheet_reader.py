"""Leitura de planilhas CSV e Excel."""

import csv
import os
from typing import List, Optional, Sequence, Tuple

import pandas as pd

from sql_generator.column_normalizer import ColumnNameNormalizer
from sql_generator.sampling import selecionar_indices_amostra


class SpreadsheetReader:
    """Lê arquivos CSV e XLSX para geração de SQL."""

    def __init__(self, codificacao: str = 'utf-8'):
        self.codificacao = codificacao

    @staticmethod
    def contar_linhas_csv(arquivo: str, codificacao: str) -> int:
        with open(arquivo, 'r', encoding=codificacao, newline='') as arquivo_csv:
            total = sum(1 for _ in arquivo_csv)
        return max(0, total - 1)

    @staticmethod
    def ler_cabecalho_csv(arquivo: str, codificacao: str) -> List[str]:
        with open(arquivo, 'r', encoding=codificacao, newline='') as arquivo_csv:
            leitor = csv.reader(arquivo_csv, delimiter=';')
            return next(leitor)

    @staticmethod
    def ler_linhas_csv_por_indices(
        arquivo: str,
        codificacao: str,
        indices: Sequence[int],
    ) -> pd.DataFrame:
        indices_desejados = set(indices)
        linhas: List[List[str]] = []
        cabecalho: List[str] = []

        with open(arquivo, 'r', encoding=codificacao, newline='') as arquivo_csv:
            leitor = csv.reader(arquivo_csv, delimiter=';')
            cabecalho = next(leitor)
            for indice, linha in enumerate(leitor):
                if indice in indices_desejados:
                    linhas.append(linha)

        return pd.DataFrame(linhas, columns=cabecalho)

    @staticmethod
    def contar_linhas_excel(arquivo: str) -> int:
        import openpyxl

        planilha = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
        try:
            aba = planilha.active
            total = 0
            for indice, _ in enumerate(aba.iter_rows(values_only=True)):
                if indice == 0:
                    continue
                total += 1
            return total
        finally:
            planilha.close()

    @staticmethod
    def ler_cabecalho_excel(arquivo: str) -> List[str]:
        import openpyxl

        planilha = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
        try:
            aba = planilha.active
            primeira_linha = next(aba.iter_rows(values_only=True))
            return [str(coluna) if coluna is not None else '' for coluna in primeira_linha]
        finally:
            planilha.close()

    @staticmethod
    def ler_linhas_excel_por_indices(
        arquivo: str,
        indices: Sequence[int],
    ) -> pd.DataFrame:
        import openpyxl

        indices_desejados = set(indices)
        cabecalho: List[str] = []
        linhas: List[List[object]] = []

        planilha = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
        try:
            aba = planilha.active
            for indice_linha, linha in enumerate(aba.iter_rows(values_only=True)):
                if indice_linha == 0:
                    cabecalho = [str(coluna) if coluna is not None else '' for coluna in linha]
                    continue
                indice_dados = indice_linha - 1
                if indice_dados in indices_desejados:
                    linhas.append(list(linha))
        finally:
            planilha.close()

        return pd.DataFrame(linhas, columns=cabecalho)

    def ler_arquivo(self, nome_arquivo: str, codificacao: Optional[str] = None) -> pd.DataFrame:
        codificacao = codificacao or self.codificacao
        extensao = os.path.splitext(nome_arquivo)[1].lower()
        if extensao == '.xlsx':
            return pd.read_excel(nome_arquivo, engine='openpyxl')
        if extensao == '.csv':
            return pd.read_csv(nome_arquivo, encoding=codificacao, delimiter=';', low_memory=False)
        raise ValueError("Formato de arquivo não suportado.")

    def ler_amostra_para_inferencia(
        self,
        nome_arquivo: str,
        codificacao: Optional[str] = None,
    ) -> Tuple[pd.DataFrame, List[str]]:
        codificacao = codificacao or self.codificacao
        extensao = os.path.splitext(nome_arquivo)[1].lower()

        if extensao == '.csv':
            total_linhas = self.contar_linhas_csv(nome_arquivo, codificacao)
            cabecalho = self.ler_cabecalho_csv(nome_arquivo, codificacao)
            indices = selecionar_indices_amostra(total_linhas)
            if not indices:
                return pd.DataFrame(columns=cabecalho), ColumnNameNormalizer.normalizar_colunas(cabecalho)
            df_amostra = self.ler_linhas_csv_por_indices(nome_arquivo, codificacao, indices)
        elif extensao == '.xlsx':
            total_linhas = self.contar_linhas_excel(nome_arquivo)
            cabecalho = self.ler_cabecalho_excel(nome_arquivo)
            indices = selecionar_indices_amostra(total_linhas)
            if not indices:
                return pd.DataFrame(columns=cabecalho), ColumnNameNormalizer.normalizar_colunas(cabecalho)
            df_amostra = self.ler_linhas_excel_por_indices(nome_arquivo, indices)
        else:
            raise ValueError("Formato de arquivo não suportado.")

        colunas = ColumnNameNormalizer.normalizar_colunas(cabecalho)
        df_amostra.columns = colunas
        return df_amostra, colunas

    @staticmethod
    def extrair_amostra_do_dataframe(df: pd.DataFrame) -> pd.DataFrame:
        """Extrai amostra distribuída de um DataFrame já carregado."""
        indices = selecionar_indices_amostra(len(df))
        if not indices:
            return df.iloc[0:0]
        return df.iloc[indices]


def ler_arquivo(nome_arquivo: str, codificacao: str = 'utf-8') -> pd.DataFrame:
    return SpreadsheetReader(codificacao).ler_arquivo(nome_arquivo, codificacao)


def ler_amostra_para_inferencia(
    nome_arquivo: str,
    codificacao: str,
) -> Tuple[pd.DataFrame, List[str]]:
    return SpreadsheetReader(codificacao).ler_amostra_para_inferencia(nome_arquivo, codificacao)


def extrair_amostra_do_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    return SpreadsheetReader.extrair_amostra_do_dataframe(df)


def contar_linhas_csv(arquivo: str, codificacao: str) -> int:
    return SpreadsheetReader.contar_linhas_csv(arquivo, codificacao)


def ler_cabecalho_csv(arquivo: str, codificacao: str) -> List[str]:
    return SpreadsheetReader.ler_cabecalho_csv(arquivo, codificacao)


def ler_linhas_csv_por_indices(
    arquivo: str,
    codificacao: str,
    indices: Sequence[int],
) -> pd.DataFrame:
    return SpreadsheetReader.ler_linhas_csv_por_indices(arquivo, codificacao, indices)


def contar_linhas_excel(arquivo: str) -> int:
    return SpreadsheetReader.contar_linhas_excel(arquivo)


def ler_cabecalho_excel(arquivo: str) -> List[str]:
    return SpreadsheetReader.ler_cabecalho_excel(arquivo)


def ler_linhas_excel_por_indices(
    arquivo: str,
    indices: Sequence[int],
) -> pd.DataFrame:
    return SpreadsheetReader.ler_linhas_excel_por_indices(arquivo, indices)
